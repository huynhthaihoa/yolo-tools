import os
import argparse
from glob import glob
from openpyxl import Workbook, load_workbook

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-a", "--ann", help="Directory contains YOLO annotation files",
                    type=str, default="coco/yolo")
    parser.add_argument("-c", "--class_file", help="Directory of class file",
                    type=str, default="obj.names")
    parser.add_argument("-d", "--csv_file", help="Directory of CSV file",
                    type=str, default="label.xlsx")
    args = parser.parse_args()
    
    csv = args.csv_file
    
    if os.path.isfile(csv) is False:
        wb = Workbook

    wb = load_workbook(csv)
    stereoSheet = wb.worksheets[0]

    classFile = open(args.class_file, "rt")
    classes = list()
    data = dict()
    for line in classFile:
        if line[-1] == '\n':
            classes.append(line[:-1])
        else:
            classes.append(line)
    
    annPaths = glob(args.ann + '/*.txt')

    for i, annpath in enumerate(annPaths):
        inFile = open(annpath, "rt")
        label = ''
        for j, line in enumerate(inFile):
            elems = line.split(' ')
            try:
                label += classes[int(elems[0])]
            except:
                continue
        stereoSheet.cell(row=(i + 1), column=1).value = annpath
        stereoSheet.cell(row=(i + 1), column=2).value = label
        print(annpath + ':' + label)
        inFile.close()
    wb.save(csv) 
    print('Generating label finished!')